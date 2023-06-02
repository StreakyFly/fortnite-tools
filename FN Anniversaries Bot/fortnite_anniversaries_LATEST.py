"""
 - Data stored in Excel: Item | Date | File name | URL
 - If there's multiple posts planned for one day, sleep 15 minutes after each post
 - Post them at around midnight - between shop sections and shop update
 - Retweet them once in the morning (when it gets posted, the tweet gets saved to a list,
        and then in the morning that tweet gets retweeted and then the next day unretweeted and removed from the list)

Copyright safety measures!
 - post video on a seperate account
 - get tweet ID
 - place it in this link: https://cdn.syndication.twimg.com/tweet?id={TWEET_ID}
 - and request data
 - get video url from the response and add it to the end of the text and tweet it
 - that's why you gotta write another function that checks if the text <= 257 (cuz 23 characters is for url)
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, date, time
from dateutil import relativedelta
from colorama import Fore, Style
import re
import requests
import os
import json
import tweepy
import time as t
import random
import traceback
import emoji
import platform


script = os.path.realpath(__file__)
if platform.system() == "Windows":
    path_list = script.split("\\")
    cwd = "\\".join(path_list[:-1]) + "\\"
elif platform.system() == "Linux":
    path_list = script.split("/")
    cwd = "/".join(path_list[:-1]) + "/"
else:
    print("PLATFORM UNDEFINED! (not in [Windows, Linux], Stopping...")
    exit()


def main():
    rc = Style.RESET_ALL

    wb_obj = openpyxl.load_workbook(f"{cwd}FN_anniversaries.xlsx")
    sheet = wb_obj.active

    data_start_row = 3
    val = Validate(rc=rc, sheet=sheet, name_column_letter='A', data_start_row=data_start_row, date_column_letter='C', video_column_letter='D', image_column_letter='E', text_time_var_column_letter='F', url_column_letter='G')

    val.scan_for_issues()  # RUN THIS WHEN UPDATING EXCEL CONTENTS TO CONFIRM EVERYTHING IS CORRECT
    # uncomment line 147 or smt like that to print changed var text when running this
    val.double_check_final()
    return

    for num, row in enumerate(sheet.iter_rows(min_row=data_start_row, max_row=sheet.max_row), start=data_start_row):
        rdate = row[2].value  # release date
        today = datetime.today()
        day_difference = (today - rdate).days
        delta = relativedelta.relativedelta(today, rdate)
        month_difference = (delta.years * 12) + delta.months

        name_id = sheet.cell(num, 1).value
        text = sheet.cell(num, 6).value
        video = sheet.cell(num, 4).value
        image = sheet.cell(num, 5).value
        if rdate.year != today.year and rdate.month == today.month and rdate.day == today.day:
            print(f"1, 2, 3... n-years anniversary: {rdate}")
            tweet_text = replace_time_var(text=text, time=delta.years, time_ymd='Y', time_vars=val._bracketify(['all', 'allN', 'allP', 'y', 'ym', 'yN', 'yP', 'ymN', 'ymP']))
            tweet(name_id=name_id, text=tweet_text, video=video, image=image)

        elif month_difference == 69 and delta.days == 0 or month_difference % 100 == 0 and delta.days == 0:
            print(f"69; 100, 200... 100*n-months old: {rdate}")
            tweet_text = replace_time_var(text=text, time=month_difference, time_ymd='M', time_vars=val._bracketify(['all', 'allN', 'allP', 'm', 'ym', 'md', 'mN', 'mP', 'ymN', 'ymP', 'mdN', 'mdP']))
            tweet(name_id=name_id, text=tweet_text, video=video, image=image)

        elif day_difference % 1000 == 0:
            print(f"1000, 2000... 1000*n-days old: {rdate}")
            tweet_text = replace_time_var(text=text, time=day_difference, time_ymd='D', time_vars=val._bracketify(['all', 'allN', 'allP', 'd', 'md', 'dN', 'dP', 'mdN', 'mdP']))
            tweet(name_id=name_id, text=tweet_text, video=video, image=image)


    # TODO: REMOVE THIS AFTER DONE TESTING!!
    # tweet_text = replace_time_var(text="The High Stakes LTM was released {ym} ago, today! #Fortnite || The High Stakes LTM is now officially {d} old! #Fortnite",
    #                  time=69, time_ymd='M', time_vars=val._bracketify(['all', 'allN', 'allP', 'd', 'md', 'dN', 'dP', 'mdN', 'mdP']))
    # tweet(name_id="High Stakes", text=tweet_text, video=None, image="fortnite_logo")
    # print(tweet_text)


def replace_time_var(text: str, time: int, time_ymd: str, time_vars: list) -> str:
    # print(f"BEFORE: {Fore.LIGHTMAGENTA_EX}{text}{Style.RESET_ALL}")
    texts = text.split(" || ")
    for t in texts:
        time_var = [i[1:-1] for i in time_vars if (i in t)]
        if time_var:
            time_var = time_var[0]
            if time_ymd == "Y":
                if str(time_var).endswith("N"):
                    pass
                elif str(time_var).endswith("P"):
                    if str(time).endswith("1") and str(time) != "11":
                        time = f"{time}st"
                    elif str(time).endswith("2") and str(time) != "12":
                        time = f"{time}nd"
                    elif str(time).endswith("3") and str(time) != "13":
                        time = f"{time}rd"
                    else:
                        time = f"{time}th"
                else:
                    if time == 1:
                        time = "1 year"
                    elif time > 1:
                        time = f"{time} years"
            elif time_ymd == "M":
                if time == 69:
                    if t.endswith("#Fortnite"):
                        t = t.replace(" #Fortnite", "")
                    t = t.strip()
                    t = t.replace("️", "")
                    while t[-1:] in emoji.EMOJI_DATA:
                        t = t[:-1]
                    if t.endswith("#Fortnite"):
                        t = t.replace("#Fortnite", "")
                    t = t + "😳" if t.endswith(" ") else t + " 😳"
                if str(time_var).endswith("N"):
                    pass
                elif str(time_var).endswith("P"):
                    if time == 69:
                        time = "69th"
                    elif time % 100 == 0:
                        time = f"{time}th"
                else:
                    time = f"{time} months"
            elif time_ymd == "D":
                if str(time_var).endswith("N"):
                    pass
                elif str(time_var).endswith("P"):
                    time = f"{time}th"
                else:
                    time = f"{time} days"

            final_text = t.replace(f"{{{time_var}}}", str(time))
            # print(f"AFTER: {Fore.LIGHTCYAN_EX}{final_text}{Style.RESET_ALL}")
            return final_text


def tweet_video(name_id: str, video):
    """
    Tweets the video on ALT account (for copyright safety) and then returns the tweeted video URL
    """
    with open(f"{cwd}video_tweets.json") as f:
        video_tweets = json.load(f)
    if name_id not in video_tweets or name_id in video_tweets and video_tweets[name_id].split(" || ")[0] != str(date.today()):
        videos = []
        for vid in video.split(" || "):
            vid_uploaded = tweepy_api_alt.media_upload(filename=f"Media/Videos/{vid}.mp4", media_category="TWEET_VIDEO")
            videos.append(vid_uploaded)
        tweet_id = tweepy_api_alt.update_status(status="", media_ids=[random.choice(videos).media_id]).id

        print(f"[{datetime.now().strftime('%d/%m/%Y-%H:%M:%S')}] [INFO] Tweeted video {tweet_id}")
        video_tweets[name_id] = f"{str(date.today())} || {tweet_id}"
        with open(f"{cwd}video_tweets.json", "w") as f:
            json.dump(video_tweets, f, indent=4)

        video_url = requests.get(url=f"https://cdn.syndication.twimg.com/tweet?id={tweet_id}").json()['entities']['media'][0]['url']
        return video_url


def tweet(name_id: str, text: str, video, image):
    """
    Example:
    run sometime after midnight
    if date matches and there are 2 posts
    check if it was already posted (if it's in the json file), otherwise post the first one and save it in the json file along with today's date
    wait 20 minutes
    check /and/ post second one and again save to json with today's date
    exit program
    crontab runs the program again after 70 minutes
    nothing
    repeat repeat repeat
    if it's between 8:40-10:40 and the post wasn't retweeted yet, it retweets it
    wait 11 minutes
    retweets next tweet if it hasn't been retweeted yet
    etc.
    if the date in the json file doesn't match current date, it unretweets that tweet and removes it from json file
    """

    with open(f"{cwd}tweets.json") as f:
        tweets = json.load(f)
    if name_id in tweets:
        un_retweet(name_id=name_id, tweets=tweets)
    else:
        media = []
        if image is not None:
            for img in image.split(" || "):
                img_uploaded = tweepy_api.media_upload(f"Media/Images/{img}.jpg")
                media.append(img_uploaded)
            tweet_id = tweepy_api.update_status(status=text, media_ids=[img.media_id for img in media]).id
        elif video is not None:
            error_continue = False
            if len(text) + 24 <= 280:  # + 24 because there must be some space between the text and the URL
                try:
                    video_url = tweet_video(name_id=name_id, video=video)
                    tweet_id = tweepy_api.update_status(status=f"{text} {video_url}").id
                except Exception as e:
                    print(f"[{datetime.now().strftime('%d/%m/%Y-%H:%M:%S')}] [ERROR] {e}")
                    error_continue = True
            if len(text) + 24 > 280 or error_continue:
                for vid in video.split(" || "):
                    vid_uploaded = tweepy_api.media_upload(filename=f"Media/Videos/{vid}.mp4", media_category="TWEET_VIDEO")
                    media.append(vid_uploaded)
                tweet_id = tweepy_api.update_status(status=text, media_ids=[random.choice(media).media_id]).id
                print(f"[{datetime.now().strftime('%d/%m/%Y-%H:%M:%S')}] [WARNING] Tweeting video on main account! {tweet_id}")

        print(f"[{datetime.now().strftime('%d/%m/%Y-%H:%M:%S')}] [INFO] Tweeted {tweet_id}")
        tweets[name_id] = f"{str(date.today())} || {tweet_id}"
        with open(f"{cwd}tweets.json", "w") as f:
            json.dump(tweets, f, indent=4)
        t.sleep(1200)


def un_retweet(name_id: str, tweets: dict):
    """
    if it's between 8:40 and 10:40 on the same day the tweet was posted and the tweet wasn't retweeted yet, it will
    retweet it
    if it's a different date compared to when the tweet was posted, it will unretweet it, if it's retweeted
    """
    if str(date.today()) == str(tweets[name_id]).split(" || ")[0] and time(hour=8, minute=40) <= datetime.now().time() <= time(hour=10, minute=40):
        tweet_id = str(tweets[name_id]).split(" || ")[1]
        tweet_status = tweepy_api.get_status(tweet_id)
        if tweet_status.retweeted is False:
            tweepy_api.retweet(tweet_id)
            print(f"[{datetime.now().strftime('%d/%m/%Y-%H:%M:%S')}] [INFO] Retweeted {tweet_id}")
            t.sleep(660)

    for name_id in list(tweets):
        if str(date.today()) != str(tweets[name_id]).split(" || ")[0]:
            tweet_id = str(tweets[name_id]).split(" || ")[1]
            tweet_status = tweepy_api.get_status(tweet_id)
            if tweet_status.retweeted is True:
                tweepy_api.unretweet(tweet_id)
                print(f"[{datetime.now().strftime('%d/%m/%Y-%H:%M:%S')}] [INFO] Unretweeted {tweet_id}")
            del tweets[name_id]
            with open(f"{cwd}tweets.json", "w") as f:
                json.dump(tweets, f, indent=4)


class Validate:
    """
    This class contains functions that check whether any data in the Excel file is missing
    or whether there are any other issues,
    such as empty cells, missing videos, images, incorrectly ordered items,
    missing time variables in the Tweet Text etc.
    """
    def __init__(self, rc, sheet, data_start_row=None, name_column_letter=None, date_column_letter=None, video_column_letter=None, image_column_letter=None, text_time_var_column_letter=None, url_column_letter=None):
        self.rc = rc
        self.sheet = sheet
        self.data_start_row = data_start_row
        self.name_column_letter = name_column_letter
        self.date_column_letter = date_column_letter
        self.video_column_letter = video_column_letter
        self.image_column_letter = image_column_letter
        self.text_time_var_column_letter = text_time_var_column_letter
        self.url_column_letter = url_column_letter

    def double_check_final(self):
        """
        improve (make more readable) this function and explain what it does here
        """
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=self.data_start_row, max_row=self.sheet.max_row), start=self.data_start_row):
            column = self.text_time_var_column_letter
            text = row[column_index_from_string(column) - 1].value
            if text is None:
                print(f"{Fore.RED}Cell empty: ({row_num}, {column}){self.rc}")
            else:
                print(text)
                replace_time_var(text=text, time=1, time_ymd='Y', time_vars=self._bracketify(['all', 'allN', 'allP', 'y', 'ym', 'yN', 'yP', 'ymN', 'ymP']))
                replace_time_var(text=text, time=2, time_ymd='Y', time_vars=self._bracketify(['all', 'allN', 'allP', 'y', 'ym', 'yN', 'yP', 'ymN', 'ymP']))
                replace_time_var(text=text, time=3, time_ymd='Y', time_vars=self._bracketify(['all', 'allN', 'allP', 'y', 'ym', 'yN', 'yP', 'ymN', 'ymP']))
                replace_time_var(text=text, time=4, time_ymd='Y', time_vars=self._bracketify(['all', 'allN', 'allP', 'y', 'ym', 'yN', 'yP', 'ymN', 'ymP']))
                replace_time_var(text=text, time=69, time_ymd='M', time_vars=self._bracketify(['all', 'allN', 'allP', 'm', 'ym', 'md', 'mN', 'mP', 'ymN', 'ymP', 'mdN', 'mdP']))
                replace_time_var(text=text, time=100, time_ymd='M', time_vars=self._bracketify(['all', 'allN', 'allP', 'm', 'ym', 'md', 'mN', 'mP', 'ymN', 'ymP', 'mdN', 'mdP']))
                replace_time_var(text=text, time=1000, time_ymd='D', time_vars=self._bracketify(['all', 'allN', 'allP', 'd', 'md', 'dN', 'dP', 'mdN', 'mdP']))

    def scan_for_issues(self):
        if self.data_start_row is None:
            print(f"{Fore.RED}PARAMETER NOT SET: 'data_start_row'{self.rc}")
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
        else:
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________{self.rc}")
            self.__check_empty()

        if self.name_column_letter is None:
            print(f"{Fore.RED}PARAMETER NOT SET: 'name_column_letter'{self.rc}")
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
        else:
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
            self.__unique_values()

        if self.date_column_letter is None:
            print(f"{Fore.RED}PARAMETER NOT SET: 'date_column_letter'{self.rc}")
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
        else:
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
            self.__check_dates()

        if self.text_time_var_column_letter is None:
            print(f"{Fore.RED}PARAMETER NOT SET: 'text_time_var_column_letter'{self.rc}")
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
        else:
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
            self.__missing_time_variable()
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
            self.__text_plus_video_url_length()

        if None in [self.video_column_letter, self.image_column_letter]:
            if self.video_column_letter is None:
                print(f"{Fore.RED}PARAMETER NOT SET: 'video_column_letter'{self.rc}")
            if self.image_column_letter is None:
                print(f"{Fore.RED}PARAMETER NOT SET: 'image_column_letter'{self.rc}")
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
        else:
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
            self.__check_media()

        if self.url_column_letter is None:
            print(f"{Fore.RED}PARAMETER NOT SET: 'url_column_letter'{self.rc}")
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________{self.rc}")
        else:
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________\n\n___________________________________________________{self.rc}")
            self.__check_url()
            print(f"{Fore.LIGHTYELLOW_EX}___________________________________________________{self.rc}")

    def __check_empty(self):
        """
        SUBFUNCTION OF scan_for_issues(sheet)
        Checks if any cell is empty (None).
        """
        print(f"{Fore.LIGHTCYAN_EX}Checking for empty cells...{self.rc}")
        filled = True
        empty_cell_no = 0
        no_of_rows = len([x for x in self.sheet.iter_rows(min_row=self.data_start_row, max_row=self.sheet.max_row)])
        for row_no in range(no_of_rows):
            for col_num, col in enumerate(self.sheet.iter_cols(min_row=self.data_start_row, max_row=self.sheet.max_row), start=1):
                value = col[row_no].value
                if value is None:
                    filled = False
                    empty_cell_no += 1
                    print(f"{Fore.RED}Cell empty: ({row_no + self.data_start_row}, {get_column_letter(col_num)}) => {value}{self.rc}")
                elif len(str(value)) < 4 and str(value) not in ['ltm', 'map']:
                    print(f"{Fore.LIGHTMAGENTA_EX}Cell value contains short text => {value}{self.rc}")

        if filled:
            print(f"{Fore.LIGHTGREEN_EX}All cells are filled.{self.rc}")
        else:
            print(f"{Fore.LIGHTRED_EX}Cells are NOT filled! ({empty_cell_no}x empty cells){self.rc}")

    def __unique_values(self):
        """
        SUBFUNCTION OF scan_for_issues(sheet)
        Checks if all values are unique.
        """
        print(f"{Fore.LIGHTCYAN_EX}Checking whether Names are unique..{self.rc}")
        items = []
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=self.data_start_row, max_row=self.sheet.max_row), start=self.data_start_row):
            column = self.name_column_letter
            text = row[column_index_from_string(column) - 1].value
            items.append(text)
        if len(items) == len(set(items)):
            print(f"{Fore.LIGHTGREEN_EX}All Names are unique.{self.rc}")
        else:
            print(f"{Fore.LIGHTRED_EX}The following Names repeat: {[x for n, x in enumerate(items) if x in items[:n]]}{self.rc}")

    def __missing_time_variable(self):
        """
        SUBFUNCTION OF scan_for_issues(sheet)
        Checks if all Tweet Text cells contain the following time variables:
                [y, m, d, all, ym, md,
                 yN, mN, dN, allN, ymN, mdN,
                 yP, mP, dP, allP, ymP, mdP]
        """
        print(f"{Fore.LIGHTCYAN_EX}Checking for missing time variables...{self.rc}")
        correct = True
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=self.data_start_row, max_row=self.sheet.max_row), start=self.data_start_row):
            column = self.text_time_var_column_letter
            text = row[column_index_from_string(column) - 1].value
            if text is None:
                print(f"{Fore.RED}Cell empty: ({row_num}, {column}){self.rc}")
                correct = False
                continue

            all_bracket_variables = re.findall("({.*?})", text)
            contains = self._contains(self._bracketify(['y', 'm', 'd', 'all', 'ym', 'md', 'yN', 'mN', 'dN', 'allN',
                                                        'ymN', 'mdN', 'yP', 'mP', 'dP', 'allP', 'ymP', 'mdP']), text)
            undefined_variables = list(set(all_bracket_variables) - set(contains))
            if undefined_variables:
                correct = False
                print(f"{Fore.LIGHTRED_EX}Cell ({row_num}, {column}) contains undefined variables: {undefined_variables}{self.rc}")

            if "{all}" in text or "{allN}" in text or "{allP}" in text:
                continue
            y_exists = m_exists = d_exists = False
            for i in contains:
                if 'y' in i:
                    y_exists = True
                if 'm' in i:
                    m_exists = True
                if 'd' in i:
                    d_exists = True
            if all([y_exists, m_exists, d_exists]):
                continue
            else:
                print(f"{Fore.LIGHTRED_EX}Cell ({row_num}, {column}) contains incorrect amount of time variables: {contains}{self.rc}")
                correct = False
        if correct:
            print(f"{Fore.LIGHTGREEN_EX}All time variables are set.{self.rc}")
        else:
            print(f"{Fore.LIGHTRED_EX}Not all time variables are set!{self.rc}")

    def __text_plus_video_url_length(self):
        """
        SUBFUNCTION OF scan_for_issues(sheet)
        Checks if Tweet Text + 24 characters (url length + 1 extra for space) <= 280 characters.
        """
        print(f"{Fore.LIGHTCYAN_EX}Checking text length with added video URL...{self.rc}")
        correct = True
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=self.data_start_row, max_row=self.sheet.max_row), start=self.data_start_row):
            column = self.text_time_var_column_letter
            text = row[column_index_from_string(column) - 1].value
            if text is None:
                print(f"{Fore.RED}Cell empty: ({row_num}, {column}){self.rc}")
                correct = False
                continue
            texts = text.split(" || ")
            for t in texts:
                if (len(t) + 24 + 7) > 280:  # + 7 is minimum extra length to replace the variable, eg. "{m}" to "100 months"
                    correct = False
                    print(f"{Fore.RED}Text POSSIBLY too long! Check final time variable length. Length: {len(t)} + 24 + 7(or however long the time var is). Cell: ({row_num}, {column})\n{Fore.MAGENTA}{t}\n{self.rc}")
        if correct:
            print(f"{Fore.LIGHTGREEN_EX}Length of all texts is short enough.{self.rc}")
        else:
            print(f"{Fore.LIGHTRED_EX}Some text lengths are possibly too long!{self.rc}")

    def __check_url(self):
        """
        SUBFUNCTION OF scan_for_issues(sheet)
        Checks if URLs exist.
        """
        print(f"{Fore.LIGHTCYAN_EX}Checking URLs...{self.rc}")
        correct = True
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=self.data_start_row, max_row=self.sheet.max_row), start=self.data_start_row):
            column = self.url_column_letter
            urls = row[column_index_from_string(column) - 1].value
            if urls is None:
                print(f"{Fore.RED}Cell empty: ({row_num}, {column}){self.rc}")
                correct = False
                continue
            for url in urls.split(" || "):
                try:
                    get = requests.get(url)
                    if get.status_code != 200:
                        correct = False
                        print(f"{Fore.RED}URL not reachable, status_code: {get.status_code}: {url}")

                except requests.exceptions.RequestException as e:
                    correct = False
                    print(f"{Fore.RED}URL not reachable: '{e}'")
        if correct:
            print(f"{Fore.LIGHTGREEN_EX}All URLs are valid.{self.rc}")
        else:
            print(f"{Fore.LIGHTRED_EX}Invalid URLs found!{self.rc}")

    def _contains(self, check_for: list, check_here: list or str) -> list:
        """y
        SUBFUNCTION OF _missing_time_variable(sheet)
        which is SUBFUNCTION OF scan_for_issues(sheet)
        Creates a list with items from check_for that are in the check_here.
        :returns a list with items from check_for in check_here - if empty, it's considered False, otherwise True
        """
        return [i for i in check_for if (i in check_here)]

    def _bracketify(self, old_list: list or str) -> list:
        """
        SUBFUNCTION OF _missing_time_variable(sheet)
        which is SUBFUNCTION OF scan_for_issues(sheet)
        Adds brackets {} to all elements in a list.
        Example:
             old_list = ['y, 'yN', 'yP']
             new_list = ['{y}', '{yN}', '{yP}']
        """
        # new_list = []
        # for i in old_list:
        #     new_list.append(f"{{{i}}}")
        # return new_list
        return [f"{{{i}}}" for i in old_list]

    def __check_media(self):
        """
        SUBFUNCTION OF scan_for_issues(sheet)
        Checks if media files exist.
        """
        print(f"{Fore.LIGHTCYAN_EX}Checking media...{self.rc}")
        missing_vid = self.__check_videos()
        missing_img = self.__check_images()
        if True in [missing_vid, missing_img]:
            print(f"{Fore.LIGHTRED_EX}Missing media files!{self.rc}")
        else:
            print(f"{Fore.LIGHTGREEN_EX}All media files exist.{self.rc}")

    def __check_videos(self):
        """
        SUBFUNCTION OF _check_media(sheet),
        which is SUBFUNCTION OF scan_for_issues(sheet)
        Checks if videos exist.
        """
        print(f"{Fore.CYAN}Videos:{self.rc}")
        missing = False
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=self.data_start_row, max_row=self.sheet.max_row), start=self.data_start_row):
            column = self.video_column_letter
            val = row[column_index_from_string(column) - 1].value
            if val is None:
                print(f"{Fore.RED}Cell empty: ({row_num}, {column}){self.rc}")
                missing = True
                continue
            videos = val.split(" || ")
            for video in videos:
                file_exists = os.path.isfile(f'Media/Videos/{video}.mp4')  # returns True if file exists, otherwise False
                if file_exists is False:
                    missing = True
                    print(f"{Fore.RED}Video '{video}.mp4' in cell ({row_num}, {column}) does NOT exist!{self.rc}")
        if missing is False:
            print(f"{Fore.LIGHTGREEN_EX}All videos exist.{self.rc}")
        else:
            print(f"{Fore.LIGHTRED_EX}Missing videos!{self.rc}")
        return missing

    def __check_images(self):
        """
        SUBFUNCTION OF _check_media(sheet),
        which is SUBFUNCTION OF scan_for_issues(sheet)
        Checks if images exist.
        """
        print(f"{Fore.CYAN}Images:{self.rc}")
        missing = False
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=self.data_start_row, max_row=self.sheet.max_row), start=self.data_start_row):
            column = self.image_column_letter
            val = row[column_index_from_string(column) - 1].value
            if val is None:
                print(f"{Fore.RED}Cell empty: ({row_num}, {column}){self.rc}")
                missing = True
                continue
            images = val.split(" || ")
            for image in images:
                file_exists = os.path.isfile(f'Media/Images/{image}.jpg')  # returns True if file exists, otherwise False
                if file_exists is False:
                    missing = True
                    print(f"{Fore.RED}Image '{image}.jpg' in cell ({row_num}, {column}) does NOT exist!{self.rc}")
        if missing is False:
            print(f"{Fore.LIGHTGREEN_EX}All images exist.{self.rc}")
        else:
            print(f"{Fore.LIGHTRED_EX}Missing images!{self.rc}")
        return missing

    def __check_dates(self):
        """
        SUBFUNCTION OF scan_for_issues(sheet)
        Checks if dates are in order.
        """
        print(f"{Fore.LIGHTCYAN_EX}Checking date order...{self.rc}")
        correct = True
        last_dd = 100000
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=self.data_start_row, max_row=self.sheet.max_row), start=self.data_start_row):
            column = self.date_column_letter
            rdate = row[column_index_from_string(column) - 1].value  # release date
            today = datetime.today()
            new_dd = (today - rdate).days  # day difference
            if last_dd < new_dd:
                correct = False
                print(f"{Fore.RED}Date order error in cell ({row_num}, {column})!{self.rc}")
            last_dd = new_dd

        if correct is True:
            print(f"{Fore.LIGHTGREEN_EX}Dates are correctly sorted.{self.rc}")
        else:
            print(f"{Fore.LIGHTRED_EX}Dates are NOT correctly sorted!{self.rc}")


if __name__ == "__main__":
    """
    Instead of infinite "While True", the script will be ran by Crontab every 70 minutes
    """
    try:
        """
        main twitter account @StreakyFly
        """
        with open(f"{cwd}TWITTER_KEYS_MAIN.json") as file:
            keysjson = json.load(file)

        ACCESS_TOKEN = keysjson["access_token"]
        ACCESS_SECRET = keysjson["access_secret"]
        CONSUMER_KEY = keysjson["consumer_key"]
        CONSUMER_SECRET = keysjson["consumer_secret"]

        authenticate = tweepy.OAuthHandler(CONSUMER_KEY, CONSUMER_SECRET)
        authenticate.set_access_token(ACCESS_TOKEN, ACCESS_SECRET)
        tweepy_api = tweepy.API(authenticate)

        """
        alt twitter account @comeupwithagoodname
        """
        with open(f"{cwd}TWITTER_KEYS_ALT.json") as file:
            keysjson_2 = json.load(file)

        ACCESS_TOKEN_2 = keysjson_2["access_token"]
        ACCESS_SECRET_2 = keysjson_2["access_secret"]
        CONSUMER_KEY_2 = keysjson_2["consumer_key"]
        CONSUMER_SECRET_2 = keysjson_2["consumer_secret"]

        authenticate_2 = tweepy.OAuthHandler(CONSUMER_KEY_2, CONSUMER_SECRET_2)
        authenticate_2.set_access_token(ACCESS_TOKEN_2, ACCESS_SECRET_2)
        tweepy_api_alt = tweepy.API(authenticate_2)

        main()
    except Exception as error:
        print(f"{Fore.RED}[{datetime.now().strftime('%d/%m/%Y-%H:%M:%S')}] [ERROR] Unknown exception occurred: {error}\n{traceback.format_exc()}{Style.RESET_ALL}")